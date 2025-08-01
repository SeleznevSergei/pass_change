import os
import string
import random
import logging
import sys
if sys.version_info >= (3, 11):
    import tomllib
else:
    import tomli as tomllib
import functools
import pandas as pd
from openpyxl import load_workbook
from datetime import date
from CGPCLI import Server


def hook_try_except(text):
    def try_except(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            try:
                result = func(*args, **kwargs)
                logging.info(f'{text} OK')
            except Exception as e:
                print(e)
                logging.error(f'{text} Failed')
                sys.exit(1)
            return result

        return wrapper

    return try_except


class Changer:
    def __init__(self, login, password, server, domain, users_xlsx, new_password):
        self.login = login
        self.password = password
        self.server = Server(server)
        self.domain = domain
        self.users_xlsx = users_xlsx
        self.new_password = new_password
        self.path = 'files'
        self.path_user_xlsx = '/'.join([self.path, self.users_xlsx])
        self.password_sheet = 'NewPasswords'
        self.action_text = None
        self.user = None
        self.user_pass_dict = {'UserName': [], 'Password': []}

    def logging_setup_and_start(self):
        current_date = date.today()
        logging.basicConfig(level=logging.DEBUG, filename=f'{self.path}/data_{current_date}.log',
                            format="%(asctime)s :: %(levelname)s :: %(message)s", datefmt='%Y-%m-%d %H:%M:%S')
        logging.info('pass_change.py started')

    def __dynamic_decorator(self, func):
        @hook_try_except(self.action_text)
        def wrapper():
            return func()
        return wrapper()

    def server_connect(self):
        self.action_text = 'Connection to server'
        self.__dynamic_decorator(lambda: self.server.connect())

    def server_login(self):
        self.action_text = 'Authorization'
        self.__dynamic_decorator(lambda: self.server.login(self.login, self.password))

    def set_password(self, sip=None):
        self.action_text = f'Changing password for {self.user}{self.domain}'
        if sip is not None:
            self.__dynamic_decorator(lambda: self.server.set_account_password(
                self.user + self.domain,
                self.new_password,
                method="method SIP"
            ))
        else:
            self.__dynamic_decorator(lambda: self.server.set_account_password(
                self.user + self.domain,
                self.new_password
            ))

    def create_excel(self, name_list: list):
        sip_lst = []
        for _ in name_list:
            sip_lst.append('')
        df = pd.DataFrame({'UserName': name_list, 'SIP_Password': sip_lst})
        try:
            df.to_excel(self.path_user_xlsx, sheet_name='DomainUsers', index=False)
            logging.info(f'{self.users_xlsx} has been created')
        except Exception as e:
            logging.error(f'Error when creating the file: {e}')

    def add_users_passwords_to_excel(self):
        df = pd.DataFrame(self.user_pass_dict)
        with pd.ExcelWriter(self.path_user_xlsx, mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name=self.password_sheet, index=False)

    def add_user_sip_passwords_to_excel(self):
        df = pd.read_excel(self.path_user_xlsx, sheet_name='DomainUsers')
        sip_dict = dict(zip(self.user_pass_dict['UserName'], self.user_pass_dict['Password']))
        df['SIP_Password'] = df['UserName'].map(sip_dict)
        df.to_excel(self.path_user_xlsx, sheet_name='DomainUsers', index=False)
        logging.info(f'SIP passwords to the {self.users_xlsx} have been added')

    def delete_xlsx_sheet(self):
        wb = load_workbook(self.path_user_xlsx)
        if self.password_sheet in wb.sheetnames:
            del wb[self.password_sheet]
            wb.save(self.path_user_xlsx)

    def gen_pass(self, length=10):
        chars = string.ascii_letters + string.digits
        self.new_password = ''.join(random.choice(chars) for _ in range(length))

    def switch_pass_off(self):
        self.action_text = f'Disable password for {self.user}{self.domain}'
        self.__dynamic_decorator(lambda: self.server.update_account_settings(self.user + self.domain, {'UseAppPassword': 'NO'}))

    def switch_pass_on(self):
        self.action_text = f'Enable password for {self.user}{self.domain}'
        self.__dynamic_decorator(lambda: self.server.update_account_settings(self.user + self.domain, {'UseAppPassword': 'YES'}))

    def get_acc_pass_on_off(self):
        self.action_text = f'Get password status for {self.user}{self.domain}'
        res = self.__dynamic_decorator(lambda: self.server.get_account_effective_settings(self.user + self.domain))
        if res['body']['UseAppPassword'].upper() == 'YES':
            return True
        return False
